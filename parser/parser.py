import os
import sys
from datetime import date, datetime

import django
import openpyxl

# for using Dajngo models
sys.path.append("api/")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "api.config.settings")
django.setup()

from api.billboards.models import Billboard, City, MonthYear, Occupation

col_title_to_billboard_field = {
    "Город": "city",
    "Тип поверхности": "surface_type",
    "Сторона": "side",
    "Адрес": "address",
    "Вн. код": "internal_key",
    "Широта": "latitude",
    "Долгота": "longitude",
    "фото/схема": "image_url",
    "Прайс C НДС": "price",
    "Диджтал кол-во показов": "number_of_displays",
    "GRP": "grp",
    "OTS": "ots",
    "Код Эспар": "aspar_code",
    "Материал носителя": "fabric",
    "Ограничения по продукту": "restrictions",
    "Городской Округ": "district",
    "Тех. требования": "technical_requirements",
    "Монтаж. Прайс  с НДС": "price_per_installation",
    "Переклейка. Прайс с НДС": "price_per_update",
    "Разрешение ПО": "permitted_until",
    "Примечание": "comment",
}

col_title_to_month_field = {
    "Июль 2023": date(2023, 7, 1),
    "Август 2023": date(2023, 8, 1),
    "Сентябрь 2023": date(2023, 9, 1),
    "Октябрь 2023": date(2023, 10, 1),
    "Ноябрь 2023": date(2023, 11, 1),
    "Декабрь 2023": date(2023, 12, 1),
}


def do_parcing(file_name: str) -> None:
    # pandas do it faster but can't read hyperlinks
    # df = pd.read_excel(file_name, sheet_name="Статус")
    # df_dicts = df.to_dict("records")

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    keys = [cell.value for cell in ws[1]]

    for row in ws[2 : ws.max_row]:
        # exclude not valid rows 28596-28605
        if row[0].internal_value:
            values = [cell.value for cell in row]
            data_dct = dict(zip(keys, values))
            data_dct["фото/схема"] = str(row[10].hyperlink.target)

            billboard_data = {}
            months_data = {}

            # rename all columns according db fields
            for key, value in data_dct.items():
                db_field = col_title_to_billboard_field.get(key, None)
                if db_field and (value is not None):
                    billboard_data[db_field] = value
                db_field = col_title_to_month_field.get(key, None)
                if db_field:
                    if value is None:
                        value = "Свободно"
                    months_data[db_field] = value

            # create City obect
            city_name = billboard_data.pop("city")
            city, _ = City.objects.get_or_create(name=city_name)
            billboard_data["city"] = city

            # create billboard object
            # some formatting
            billboard_data["has_backlight"] = data_dct["Осв"] == "Да"

            permitted_until = billboard_data.get("permitted_until", None)
            if permitted_until:
                billboard_data["permitted_until"] = datetime.strptime(
                    permitted_until, "%d.%m.%Y"
                ).date()
            else:
                billboard_data.pop("permitted_until")

            billboard = Billboard.objects.create(**billboard_data)

            # create occupation table
            for month_id in range(7, 13):
                month, _ = MonthYear.objects.get_or_create(month=date(2023, month_id, 1))

                state = months_data[month.month].strip()
                comment = ""

                if state == "Свободно":
                    state = Occupation.FREE
                elif state == "Продано":
                    state = Occupation.SOLD
                elif "Продано" in state:
                    state = Occupation.PARTLY
                elif "Не установлена" in state:
                    state = Occupation.NOT_INSTALLED
                elif "Зарезервирована" in state:
                    state = Occupation.RESERVED
                else:
                    comment = state
                    state = Occupation.OTHER

            Occupation.objects.create(
                billboard=billboard, month=month, state=state, comment=comment
            )
